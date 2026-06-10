"""Gera CSV municipal de comunidades quilombolas certificadas.

Uso:
    uv run --with openpyxl python reports/ATM-EQ/scripts/build_quilombola_certified_by_municipio.py
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from google.cloud import bigquery
from openpyxl import load_workbook

# Permite importar o pacote local sem instalar em modo editable.
# script: school-report-python/reports/ATM-EQ/scripts/build_quilombola_certified_by_municipio.py
# root:   school-report-python/
ROOT_DIR = Path(__file__).resolve().parents[3]
SRC_DIR = ROOT_DIR / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from schoolreport.config import get_local_settings  # noqa: E402

DEFAULT_INPUT = (
    ROOT_DIR / "reports" / "ATM-EQ" / "docs" / "ComunidadesQuilombolas FCP.xlsx"
)
DEFAULT_OUTPUT = (
    ROOT_DIR
    / "output"
    / "atm-eq"
    / "comunidades_quilombolas_certificadas_municipio.csv"
)


@dataclass
class SourceRow:
    uf_raw: str
    municipio_raw: str
    ibge_raw: str
    comunidade_raw: str
    processo_certificacao: str
    data_certificacao: date | None
    situacao_certificacao: str


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", "\n")
    text = re.sub(r"\s*\n\s*", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def normalize_status(value: object) -> str:
    return normalize_text(value).upper()


def normalize_lookup_text(value: object) -> str:
    text = normalize_text(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return text.upper()


def parse_excel_date(value: object) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = normalize_text(value)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def extract_ibge_codes(value: object) -> list[int]:
    text = str(value or "")
    codes = re.findall(r"\d{7}", text)
    ordered_unique: list[int] = []
    seen: set[int] = set()
    for code in codes:
        as_int = int(code)
        if as_int not in seen:
            seen.add(as_int)
            ordered_unique.append(as_int)
    return ordered_unique


def load_source_rows(path: Path) -> list[SourceRow]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    rows: list[SourceRow] = []
    first = True
    for row in sheet.iter_rows(values_only=True):
        if first:
            first = False
            continue

        if row is None:
            continue

        rows.append(
            SourceRow(
                uf_raw=normalize_text(row[0]),
                municipio_raw=normalize_text(row[1]),
                ibge_raw=normalize_text(row[2]),
                comunidade_raw=normalize_text(row[3]),
                processo_certificacao=normalize_text(row[4]),
                data_certificacao=parse_excel_date(row[5]),
                situacao_certificacao=normalize_status(row[6]),
            )
        )

    return rows


def fetch_municipio_lookup(
    project_id: str,
) -> tuple[dict[int, dict[str, object]], dict[tuple[str, str], int]]:
    client = bigquery.Client(project=project_id)
    query = """
    SELECT
      CAST(id_municipio AS INT64) AS id_municipio,
      nome AS municipio,
      sigla_uf
    FROM `br-mec-segape.educacao_dados_mestres.municipio`
    """
    rows = client.query(query).result()
    by_id = {
        int(row.id_municipio): {
            "municipio": row.municipio,
            "sigla_uf": row.sigla_uf,
        }
        for row in rows
    }
    by_name = {
        (
            normalize_lookup_text(payload["sigla_uf"]),
            normalize_lookup_text(payload["municipio"]),
        ): id_municipio
        for id_municipio, payload in by_id.items()
    }
    return by_id, by_name


def build_output_rows(
    source_rows: list[SourceRow],
    municipio_lookup: dict[int, dict[str, object]],
    municipio_name_lookup: dict[tuple[str, str], int],
    data_ultima_atualizacao: date,
) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    output_rows: list[dict[str, object]] = []
    rejected_rows: list[dict[str, object]] = []
    seen_keys: set[tuple[int, str, str]] = set()

    for source in source_rows:
        if source.situacao_certificacao != "CERTIFICADA":
            continue

        ibge_codes = extract_ibge_codes(source.ibge_raw)
        if not ibge_codes:
            rejected_rows.append(
                {
                    "motivo_rejeicao": "sem_ibge_valido",
                    "uf_raw": source.uf_raw,
                    "municipio_raw": source.municipio_raw,
                    "ibge_raw": source.ibge_raw,
                    "comunidade_raw": source.comunidade_raw,
                    "processo_certificacao": source.processo_certificacao,
                    "data_certificacao": (
                        source.data_certificacao.isoformat()
                        if source.data_certificacao
                        else ""
                    ),
                    "situacao_certificacao": source.situacao_certificacao,
                }
            )
            continue

        for ibge_code in ibge_codes:
            municipio = municipio_lookup.get(ibge_code)
            if municipio is None:
                fallback_key = (
                    normalize_lookup_text(source.uf_raw),
                    normalize_lookup_text(source.municipio_raw),
                )
                fallback_id = municipio_name_lookup.get(fallback_key)
                if fallback_id is not None:
                    ibge_code = fallback_id
                    municipio = municipio_lookup[fallback_id]

            if municipio is None:
                rejected_rows.append(
                    {
                        "motivo_rejeicao": "ibge_nao_encontrado_no_cadastro_mestre",
                        "uf_raw": source.uf_raw,
                        "municipio_raw": source.municipio_raw,
                        "ibge_raw": source.ibge_raw,
                        "comunidade_raw": source.comunidade_raw,
                        "processo_certificacao": source.processo_certificacao,
                        "data_certificacao": (
                            source.data_certificacao.isoformat()
                            if source.data_certificacao
                            else ""
                        ),
                        "situacao_certificacao": source.situacao_certificacao,
                        "id_municipio_tentado": ibge_code,
                    }
                )
                continue

            key = (
                ibge_code,
                source.comunidade_raw,
                source.processo_certificacao,
            )
            if key in seen_keys:
                continue
            seen_keys.add(key)

            output_rows.append(
                {
                    "sigla_uf": municipio["sigla_uf"],
                    "id_municipio": ibge_code,
                    "municipio": municipio["municipio"],
                    "comunidade": source.comunidade_raw,
                    "processo_certificacao": source.processo_certificacao,
                    "data_certificacao": (
                        source.data_certificacao.isoformat()
                        if source.data_certificacao
                        else ""
                    ),
                    "situacao_certificacao": source.situacao_certificacao,
                    "data_ultima_atualizacao": data_ultima_atualizacao.isoformat(),
                }
            )

    output_rows.sort(key=lambda row: (row["sigla_uf"], row["id_municipio"], row["comunidade"]))
    rejected_rows.sort(
        key=lambda row: (
            row.get("motivo_rejeicao", ""),
            row.get("ibge_raw", ""),
            row.get("comunidade_raw", ""),
        )
    )
    return output_rows, rejected_rows


def write_csv(path: Path, rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return

    fieldnames = list(rows[0].keys())
    with path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Gera CSV de comunidades quilombolas certificadas por município."
    )
    parser.add_argument(
        "--input",
        default=str(DEFAULT_INPUT),
        help="Caminho do arquivo XLSX de origem.",
    )
    parser.add_argument(
        "--output",
        default=str(DEFAULT_OUTPUT),
        help="Caminho do CSV final de saída.",
    )
    parser.add_argument(
        "--project-id",
        default=None,
        help="Projeto GCP. Se omitido, usa GCP_PROJECT_ID do .env.",
    )
    parser.add_argument(
        "--data-ultima-atualizacao",
        default=date.today().isoformat(),
        help="Data de atualização no formato YYYY-MM-DD.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    settings = get_local_settings()
    project_id = args.project_id or settings.gcp_project_id
    if not project_id:
        print("Projeto nao definido. Configure GCP_PROJECT_ID no .env.")
        return 2

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"Arquivo de entrada nao encontrado: {input_path}")
        return 1

    try:
        data_ultima_atualizacao = datetime.strptime(
            args.data_ultima_atualizacao, "%Y-%m-%d"
        ).date()
    except ValueError:
        print("Parametro --data-ultima-atualizacao invalido. Use YYYY-MM-DD.")
        return 1

    source_rows = load_source_rows(input_path)
    municipio_lookup, municipio_name_lookup = fetch_municipio_lookup(project_id)
    output_rows, rejected_rows = build_output_rows(
        source_rows=source_rows,
        municipio_lookup=municipio_lookup,
        municipio_name_lookup=municipio_name_lookup,
        data_ultima_atualizacao=data_ultima_atualizacao,
    )

    output_path = Path(args.output)
    rejected_path = output_path.with_name(output_path.stem + "_rejeitados.csv")
    write_csv(output_path, output_rows)
    write_csv(rejected_path, rejected_rows)

    certified_source_rows = sum(
        1 for row in source_rows if row.situacao_certificacao == "CERTIFICADA"
    )
    print(f"Arquivo de entrada: {input_path}")
    print(f"CSV final: {output_path}")
    print(f"CSV rejeitados: {rejected_path}")
    print(f"Linhas na planilha: {len(source_rows)}")
    print(f"Linhas certificadas na fonte: {certified_source_rows}")
    print(f"Linhas finais no CSV municipal: {len(output_rows)}")
    print(f"Linhas rejeitadas: {len(rejected_rows)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
